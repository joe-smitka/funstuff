import msoffcrypto
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import io

def decrypt_excel(file_path, password):
    decrypted = io.BytesIO()
    with open(file_path, "rb") as f:
        office_file = msoffcrypto.OfficeFile(f)
        office_file.load_key(password=password)
        office_file.decrypt(decrypted)
    decrypted.seek(0)
    return decrypted

# === SETTINGS ===
password = "yourpassword"
file1_path = "file1.xlsx"
file2_path = "file2.xlsx"
output_path = "highlighted_changes.xlsx"
id_column = "ID"  # Use the header name, e.g., 'ID'

# === STEP 1: Decrypt both files ===
decrypted_file1 = decrypt_excel(file1_path, password)
decrypted_file2 = decrypt_excel(file2_path, password)

# === STEP 2: Load into pandas ===
df1 = pd.read_excel(decrypted_file1)
df2 = pd.read_excel(decrypted_file2)

# === STEP 3: Align using ID ===
df1.set_index(id_column, inplace=True)
df2.set_index(id_column, inplace=True)

# Debugging: Check for duplicate IDs
dupes1 = df1.index[df1.index.duplicated()].unique()
dupes2 = df2.index[df2.index.duplicated()].unique()

if not dupes1.empty:
    print("Duplicate IDs in file1.xlsx:", dupes1.tolist())
if not dupes2.empty:
    print("Duplicate IDs in file2.xlsx:", dupes2.tolist())


# Include all IDs from both files
all_ids = df1.index.union(df2.index)
df1 = df1.reindex(all_ids)
df2 = df2.reindex(all_ids)

# Null-safe compare
df1_filled = df1.fillna("__NA__")
df2_filled = df2.fillna("__NA__")
diff_df = df1_filled.ne(df2_filled)

# === STEP 4: Reload file2 into openpyxl ===
decrypted_file2.seek(0)
wb = load_workbook(filename=decrypted_file2)
ws = wb.active

# === STEP 5: Get actual column number for ID ===
id_col_num = None
for col in range(1, ws.max_column + 1):
    if ws.cell(row=1, column=col).value == id_column:
        id_col_num = col
        break
if id_col_num is None:
    raise ValueError(f"Couldn't find column header: '{id_column}'")

# === STEP 6: Build ID to row mapping ===
id_to_row = {}
for row in range(2, ws.max_row + 1):  # skip header
    val = ws.cell(row=row, column=id_col_num).value
    if val is not None:
        id_to_row[val] = row

# === STEP 7: Build column name to Excel column number mapping ===
col_name_to_num = {}
for col in range(1, ws.max_column + 1):
    header = ws.cell(row=1, column=col).value
    if header:
        col_name_to_num[header] = col

# === STEP 8: Highlight differences ===
highlight_change = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
highlight_new = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

for idx, row in diff_df.iterrows():
    if idx not in id_to_row:
        continue
    excel_row = id_to_row[idx]
    for col_name, changed in row.items():
        if changed:
            col_num = col_name_to_num.get(col_name)
            if col_num:
                ws.cell(row=excel_row, column=col_num).fill = highlight_change

# === STEP 9: Highlight brand new rows (only in df2) ===
new_ids = df2.index.difference(df1.index)
for new_id in new_ids:
    if new_id not in id_to_row:
        continue
    row_num = id_to_row[new_id]
    for col in range(1, ws.max_column + 1):
        ws.cell(row=row_num, column=col).fill = highlight_new

# === STEP 10: Save the result ===
wb.save(output_path)
print(f"Done! Changes saved to '{output_path}'")
