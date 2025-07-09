import msoffcrypto
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import io

def decrypt_excel(file_path, password):
    decrypted = io.BytesIO()
    with open(file_path, "rb") as f:
        office_file = msoffcrypto.OfficeFile(f)
        office_file.load_key(password=password)
        office_file.decrypt(decrypted)
    decrypted.seek(0)
    return decrypted

# === SETTINGS ===
password = "yourpassword"
file1_path = "file1.xlsx"
file2_path = "file2.xlsx"
output_path = "highlighted_changes.xlsx"

# === STEP 1: Decrypt both files ===
decrypted_file1 = decrypt_excel(file1_path, password)
decrypted_file2 = decrypt_excel(file2_path, password)

# === STEP 2: Read both into pandas DataFrames ===
df1 = pd.read_excel(decrypted_file1)
df2 = pd.read_excel(decrypted_file2)

# === STEP 3: Compare the two DataFrames ===
diff_df = df1.ne(df2)

# === STEP 4: Reload decrypted_file2 into openpyxl for highlighting ===
decrypted_file2.seek(0)  # rewind after pandas read
wb = load_workbook(filename=decrypted_file2)
ws = wb.active

# Yellow fill for changes
highlight = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# === STEP 5: Highlight differences ===
for row in range(2, len(df2) + 2):  # +2 for 1-based index and header row
    for col in range(1, len(df2.columns) + 1):
        if diff_df.iloc[row - 2, col - 1]:
            ws.cell(row=row, column=col).fill = highlight

# === STEP 6: Save the result ===
wb.save(output_path)
print(f"Done! Highlighted differences saved to {output_path}")
