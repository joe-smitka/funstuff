import msoffcrypto
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
import io

def decrypt_excel(file_path, password):
    decrypted = io.BytesIO()
    with open(file_path, "rb") as f:
        office_file = msoffcrypto.OfficeFile(f)
        office_file.load_key(password=password)
        office_file.decrypt(decrypted)
    decrypted.seek(0)
    return decrypted

# === SETTINGS ===
password = "yourpassword"
file1_path = "file1.xlsx"
file2_path = "file2.xlsx"
output_path = "highlighted_changes.xlsx"
id_column = "ID"
summary_column_name = "Added/Changed"

# === STEP 1: Decrypt both files ===
decrypted_file1 = decrypt_excel(file1_path, password)
decrypted_file2 = decrypt_excel(file2_path, password)

# === STEP 2: Load into pandas ===
df1 = pd.read_excel(decrypted_file1)
df2 = pd.read_excel(decrypted_file2)

# === STEP 3: Set index and deduplicate ===
df1.set_index(id_column, inplace=True)
df2.set_index(id_column, inplace=True)
df1 = df1[~df1.index.duplicated(keep='last')]
df2 = df2[~df2.index.duplicated(keep='last')]

# Reindex both to align all IDs
all_ids = df1.index.union(df2.index)
df1 = df1.reindex(all_ids)
df2 = df2.reindex(all_ids)

# === STEP 4: Compare with null-safe fill ===
df1_filled = df1.fillna("__NA__")
df2_filled = df2.fillna("__NA__")
diff_df = df1_filled.ne(df2_filled)

# === STEP 5: Reload updated Excel for highlighting ===
decrypted_file2.seek(0)
wb = load_workbook(filename=decrypted_file2)
ws = wb.active

# === STEP 6: Header map ===
header_row = 1
headers = {ws.cell(row=header_row, column=col).value: col for col in range(1, ws.max_column + 1)}

if id_column not in headers:
    raise ValueError(f"Missing ID column: {id_column}")
if summary_column_name not in headers:
    raise ValueError(f"Missing summary column: {summary_column_name}")

id_col_num = headers[id_column]
summary_col_num = headers[summary_column_name]

# === STEP 7: Map ID to row number ===
id_to_row = {}
for row in range(2, ws.max_row + 1):
    val = ws.cell(row=row, column=id_col_num).value
    if val is not None:
        id_to_row[val] = row

# === STEP 8: Highlighting styles ===
highlight_change = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
highlight_new = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
wrap_alignment = Alignment(wrap_text=True)

# === STEP 9: Apply diffs and summarize ===
for idx in diff_df.index:
    excel_row = id_to_row.get(idx)
    summary_lines = []

    is_new_row = idx in df2.index and (idx not in df1.index or df1.loc[idx].isnull().all())

    if is_new_row:
        if excel_row:
            for col in range(1, ws.max_column + 1):
                ws.cell(row=excel_row, column=col).fill = highlight_new
            summary_lines.append("Row Added")
    else:
        # Compare changed values
        row_diff = diff_df.loc[idx]
        for col_name, changed in row_diff.items():
            if changed and col_name in headers:
                col_num = headers[col_name]
                if excel_row:
                    ws.cell(row=excel_row, column=col_num).fill = highlight_change
                    old_val = df1.at[idx, col_name]
                    new_val = df2.at[idx, col_name]
                    summary_lines.append(f"{col_name}: '{old_val}' → '{new_val}'")

    if excel_row and summary_lines:
        cell = ws.cell(row=excel_row, column=summary_col_num)
        cell.value = "\n".join(summary_lines)
        cell.alignment = wrap_alignment

# === STEP 10: Save it
wb.save(output_path)
print(f"✅ Done! Saved to: {output_path}")
